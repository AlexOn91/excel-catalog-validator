
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import warnings
warnings.filterwarnings(
    "ignore",
    message="Unknown extension is not supported and will be removed",
    category=UserWarning,
    module="openpyxl\\.worksheet\\._reader"
)

import tkinter as tk
from tkinter import messagebox
import os
from openpyxl import Workbook

class ReportDownloader:
    # …

    def save_report(self, fails_for_check, save_path):
        """
        Încearcă să salveze raportul în save_path.
        Dacă fișierul e deschis în Excel, afișează un mesaj și salvează sub un nume alternativ.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Format Fails"
        ws.append(["Check Name", "Example", "Cell Reference"])

        for name, example, ref in fails_for_check:
            ws.append([name, example, ref])

        try:
            wb.save(save_path)
            return save_path
        except PermissionError:
            # Fişierul este blocat de Excel: notificăm utilizatorul
            root = tk.Tk()
            root.withdraw()
            messagebox.showwarning(
                "Raport blocat",
                f"Nu se poate salva raportul deoarece:\n'{save_path}' este deschis în Excel.\n"
                "Închideţi-l sau salvaţi manual cu un alt nume."
            )
            root.destroy()
            # Salvăm sub un nume alternativ
            base, ext = os.path.splitext(save_path)
            alt_path = f"{base}_failreport{ext}"
            wb.save(alt_path)
            return alt_path



def export_data_format_fails(report_data, df, wb_original, mapping, save_path, sheet_name):
    """
    Exportează toate erorile (❌ Fail) din grupa "Data Format Checks" într-un fișier Excel.

    Parametri:
    - report_data: dict-ul complet cu raportul de validări
    - df: pandas.DataFrame citit din Excel
    - wb_original: openpyxl.Workbook încărcat din fișierul original
    - mapping: dict {prop: header} folosit la group_b
    - save_path: calea completă unde se salvează fișierul .xlsx
    """
    ws = wb_original[sheet_name]
    checks = [
        ("Demo Data", re.compile(r'\bdemo(?:brand|sku|_category)?\b', re.IGNORECASE)),
        ("Special Characters", set("©$€£¥™®@")),
        ("Formulas", "="),
        ("HTML Tags", re.compile(r'<[^>]+>|&lt;[^&]+&gt;'))
    ]

    # aflăm care verificări au fost cu Fail în raportul inițial
    failed = {
        it["Check Performed"]
        for it in report_data.get("Data Format Checks", [])
        if it.get("Check Outcome", "").startswith("❌")
}
# păstrăm ordinea originală, dar doar pe cele fail‑uite
    checks = [
        (name, pattern)
        for name, pattern in [
            ("Demo Data",           re.compile(r'\bdemo(?:brand|sku|_category)?\b', re.IGNORECASE)),
            ("Special Characters",  set("©$€£¥™®@")),
            ("Formulas",            "="),
            ("HTML Tags",           re.compile(r'<[^>]+>|&lt;[^&]+&gt;'))
        ]
        if name in failed
]


    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    sheet = None
    for name, pattern in checks:
        fails_for_check = []
        for prop, header in mapping.items():
            if header not in df.columns:
                continue
            # găsește litera coloanei în wb_original
            col_letter = None
            for cell in ws[1]:
                if str(cell.value).strip() == header:
                    col_letter = cell.column_letter
                    break
            if not col_letter:
                continue

            for idx, raw in enumerate(ws[col_letter][1:], start=2):
                text = str(raw.value or "").strip()

                if name == "Demo Data":
        # —> un entry pentru fiecare “demo”
                    for m in pattern.finditer(text):
                        fails_for_check.append((idx, m.group(0), f"{col_letter}{idx}"))

                elif name == "Special Characters":
        # păstrează cum ai deja, per‑occurrence
                    counts = {ch: text.count(ch) for ch in pattern if text.count(ch) > 0}
                    for ch, cnt in counts.items():
                        for _ in range(cnt):
                            fails_for_check.append((idx, ch, f"{col_letter}{idx}"))
                # elif name == "Formulas":
                # # —> un entry pentru fiecare “=” din text
                #     eq_count = text.count("=")
                #     for _ in range(eq_count):
                #         fails_for_check.append((idx, "=", f"{col_letter}{idx}"))
                elif name == "Formulas":
    # un singur entry dacă celula începe cu “=”
                    if text.startswith("="):
                        fails_for_check.append((idx, "=", f"{col_letter}{idx}"))


                elif name == "HTML Tags":
        # —> un entry pentru fiecare tag găsit
                    for tag in pattern.findall(text):
                        fails_for_check.append((idx, tag, f"{col_letter}{idx}"))


                if not fails_for_check:
                    continue

        # creează sheet pentru acest check
        sheet = wb_out.create_sheet(title=name[:31])
        sheet.append(["Check Performed", "Explanation", "Cell Fail Reference"])
        for _row, example, ref in fails_for_check:
            explanation = {
                "Demo Data": f"Demo keyword '{example}' found",
                "Special Characters": f"Special character '{example}' found",
                "Formulas": f"Formula detected '{example}' found",
                "HTML Tags": f"HTML tag '{example}' found"
            }[name]
            sheet.append([name, explanation, ref])

    wb_out.save(save_path)
    return save_path



if __name__ == '__main__':
    import sys
    import pandas as pd
    from openpyxl import load_workbook

    if len(sys.argv) != 6:
        # print("Usage: python export_fail_report.py <report_json> <excel_path> <mapping_json> <sheet_name> <output_path>")
        sys.exit(1)

    import json
    report_data = json.loads(sys.argv[1])
    excel_file  = sys.argv[2]
    mapping     = json.loads(sys.argv[3])
    sheet_name  = sys.argv[4]
    output_path = sys.argv[5]

    # încarcă datele
    df = pd.read_excel(excel_file, sheet_name=sheet_name, dtype=str)
    wb_original = load_workbook(excel_file, data_only=True)

    path = export_data_format_fails(report_data, df, wb_original, mapping, output_path)
    # print(f"Fail report saved to {path}")
