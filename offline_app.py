import tkinter as tk
from tkinter import ttk, filedialog, messagebox
# import pandas as pd
import json
# from openpyxl import load_workbook
from validator import validate_file
from downloadfailreport import export_data_format_fails

APP_VERSION = "v1.0"

# import builtins

# DEBUG = False  # pune True când vrei să vezi din nou DEBUG‑urile

# # păstrăm referinţa la print-ul original
# _original_print = builtins.print

# if not DEBUG:
#     # suprascriem print în acest modul: devine no‑op
#     def print(*args, **kwargs):
#         pass
# else:
#     # readucem print-ul original
#     print = _original_print



class ToolTip:
    """Un tooltip simplu pentru orice widget Tkinter."""
    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None

    def show(self, text, x, y):
        # ascunde ce era înainte
        self.hide()
        if not text:
            return

        # fereastra fără decor
        tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw, text=text, justify="left",
            background="#ffffe0", relief="solid", borderwidth=1,
            wraplength=400
        )
        label.pack(ipadx=1, ipady=1)
        self.tipwindow = tw

    def hide(self):
        if self.tipwindow:
            self.tipwindow.destroy()
            self.tipwindow = None



# Utility pentru frame scrollabil
class ScrollableFrame(ttk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        canvas = tk.Canvas(self)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        scrollable = ttk.Frame(canvas)
        scrollable.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.scrollable_frame = scrollable

# Tab pentru maparea coloanelor
class MappingTab(ttk.Frame):
    def __init__(self, parent, file_cols, expected_props, on_validate_callback):
 
        
        super().__init__(parent)
        self.file_cols = file_cols
        self.expected = expected_props
        self.on_validate_callback = on_validate_callback
        self.mapping_vars = {}
        self.codes_rows = []

        # Scrollable
        canvas = tk.Canvas(self)
        vsb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        container = ttk.Frame(canvas)
        container.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=container, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # 1) Mapping Details
        lf = ttk.Labelframe(container, text="Mapping Details", padding=10)
        lf.grid(row=0, column=0, sticky="ew", padx=10, pady=5)
        ttk.Label(lf, text="Property", font=(None,10,'bold'))\
            .grid(row=0, column=0, sticky="w")
        ttk.Label(lf, text="Map & Rename Columns", font=(None,10,'bold'))\
            .grid(row=0, column=1, sticky="w", padx=20)

        # deasupra buclei, definește lista cu câmpurile obligatorii
        required = ["Product Name (Local Language)","Product ID", "Brand"]

        for i, prop in enumerate(self.expected, start=1):
    # punem * doar dacă proprietatea e exact Product Name (English)
            label = prop + (" *" if prop in required else "")
            ttk.Label(lf, text=label)\
                .grid(row=i, column=0, sticky="w", pady=2)
            var = tk.StringVar()
            cb = ttk.Combobox(
                lf, textvariable=var,
                values=[""] + self.file_cols,
                state="readonly", width=30)
   
            cb.grid(row=i, column=1, padx=5, pady=2)
            self.mapping_vars[prop] = var

        # 2) Additional Product Codes
        self.codes_frame = ttk.Labelframe(
            container, text="Additional Product Codes and Metadata", padding=10
        )
        self.codes_frame.grid(
            row=1, column=0, sticky="ew", padx=10, pady=5
        )
        ttk.Label(self.codes_frame, text="Code Type", font=(None,10,'bold'))\
            .grid(row=0, column=0, sticky="w")
        ttk.Label(self.codes_frame, text="Mapped Column", font=(None,10,'bold'))\
            .grid(row=0, column=1, sticky="w", padx=20)
        ttk.Button(
            self.codes_frame,
            text="+ Add additional code",
            command=self._add_code
        ).grid(row=0, column=2, sticky="w")

        # 3) Validate button
        ttk.Button(
            container,
            text="Validate with Mapping",
            command=self._on_validate
        ).grid(row=2, column=0, pady=15)

    def _add_code(self):
        i = len(self.codes_rows) + 1
        code_var = tk.StringVar()
        custom_var = tk.StringVar()
        map_var = tk.StringVar()

        # Code Type
        cb1 = ttk.Combobox(
            self.codes_frame,
            textvariable=code_var,
            values=["", "SKU", "EAN", "UPC", "GTIN", "CTIN", "ASIN", "Other"],
            state="readonly", width=15
        )
        cb1.grid(row=i, column=0, pady=2, sticky="w")

        # Mapped Column
        cb2 = ttk.Combobox(
            self.codes_frame,
            textvariable=map_var,
            values=[""] + self.file_cols,
            state="readonly", width=30
        )
        cb2.grid(row=i, column=1, padx=5, pady=2, sticky="w")

        # Custom entry, hidden
        entry_custom = ttk.Entry(
            self.codes_frame, textvariable=custom_var, width=15
        )
        entry_custom.grid(row=i, column=2, pady=2, sticky="w")
        entry_custom.grid_remove()

        def on_type(event):
            if code_var.get() == "Other":
                entry_custom.grid()
            else:
                entry_custom.grid_remove()
                custom_var.set("")
        cb1.bind("<<ComboboxSelected>>", on_type)

        # Delete button
        btn_del = ttk.Button(
            self.codes_frame,
            text="-",
            command=lambda idx=i-1: self._remove_code(idx)
        )
        btn_del.grid(row=i, column=3, padx=5, pady=2)

        self.codes_rows.append(
            (code_var, custom_var, map_var, cb1, cb2, entry_custom, btn_del)
        )

    def _remove_code(self, idx):
        # rebuild
        for w in self.codes_frame.winfo_children():
            w.destroy()
        ttk.Label(self.codes_frame, text="Code Type", font=(None,10,'bold'))\
            .grid(row=0, column=0, sticky="w")
        ttk.Label(self.codes_frame, text="Mapped Column", font=(None,10,'bold'))\
            .grid(row=0, column=1, sticky="w", padx=20)
        ttk.Button(
            self.codes_frame,
            text="+ Add additional code",
            command=self._add_code
        ).grid(row=0, column=2, sticky="w")

        rows = list(self.codes_rows)
        self.codes_rows.clear()
        del rows[idx]
        for code_var, custom_var, map_var, *_ in rows:
            self._add_code()
            last = self.codes_rows[-1]
            last[0].set(code_var.get())
            last[1].set(custom_var.get())
            last[2].set(map_var.get())

    def _on_validate(self):
    # --- 0) Warning dacă n-ai mapat niciun câmp de bază și niciun extra code ---
        no_base   = not any(var.get().strip() for var in self.mapping_vars.values())
        no_extras = not any(map_var.get().strip() for _,_, map_var,*_ in self.codes_rows)
        if no_base and no_extras:
            messagebox.showwarning(
            "Mapare necesară",
            "Trebuie să mapezi cel puțin un câmp înainte de validare."
        )
            return
    
        to_rename   = {}
        extra_props = []

    # 1a) Mapare proprietăți de bază
        for prop, var in self.mapping_vars.items():
            col = var.get().strip()
            if col:
                to_rename[col] = prop

    # 1b) Mapare coduri suplimentare (inclusiv “Other”)
        for code_var, custom_var, map_var, *_ in self.codes_rows:
            code_type = code_var.get().strip()   # ex. "EAN" sau "Other"
            col       = map_var.get().strip()    # header-ul din fișier
            if not code_type or not col:
                continue

            if code_type == "Other":
            # dacă nu ai introdus nimic în custom, fallback la numele coloanei
                prop = custom_var.get().strip() or col
            else:
                prop = code_type

            to_rename[col]   = prop
            extra_props.append(prop)

    # --- 2) Verifică că ai măcar o coloană mapată ---
        if not to_rename:
            messagebox.showerror(
            "Mapping Error",
            "Mapările nu au fost citite corect; încearcă din nou."
        )
            return

    # --- 3) Construiește mapped_props ca prop_name → column_name ---
        mapped_props = { prop: col for col, prop in to_rename.items() }

    # --- 4) Apelează callback-ul principal ---
        self.on_validate_callback(to_rename, extra_props, mapped_props)

# Aplicația principală
class OfflineCatalogValidatorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.loaded_path = None
        self.selected_sheet = None
        self.df = None
        self.wb = None
        self.cols = []

        self.title(f"Offline Catalog Validator {APP_VERSION}")
        self.geometry("900x700")

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True)

         # inițializări pentru export
        self.current_report = {}
        self.current_mapping = {}

        # Tab 1: Load File
        self.tab1 = ttk.Frame(self.nb)
        self.nb.add(self.tab1, text="Load File")
        ttk.Button(
            self.tab1,
            text="Select xlsx Excel File…",
            command=self.on_load
        ).pack(pady=20)
        # ===== inserție începe aici =====
# frame care va conține Combobox-ul pentru sheet-uri
        self.sheet_frame = ttk.Frame(self.tab1)
# label + combobox setate, dar ascunse la start
        ttk.Label(self.sheet_frame, text="Select sheet:").pack(side="left", padx=(0,5))
        self.sheet_var = tk.StringVar()
        self.sheet_cb  = ttk.Combobox(
        self.sheet_frame,
        textvariable=self.sheet_var,
        state="readonly",
        values=[],
        width=50
        )
        
        self.sheet_cb.pack(side="left", padx=(0,5))
        self.sheet_cb.set("")   # dropdown-ul începe gol

        # Adaugă binding-ul
        self.sheet_cb.bind("<<ComboboxSelected>>", self._on_sheet_selected)
# ascunde frame-ul până când e nevoie
        self.sheet_frame.pack_forget()
        
        # 2) Sub tot conținutul (în aceeași indentare cu self.nb), adaugă status bar‑ul:
        ttk.Label(
            self.tab1,
            text="Created by Alex Oniciuc",
            anchor="e",
            background=	"#3E0E8C",
            foreground="white",
            font=(None, 8, "italic")
        ).place(relx=1.0, rely=1.0, anchor="se", x=-5, y=-5)
       
        # Tab 2: Map Columns
        self.tab2 = ttk.Frame(self.nb)
        self.nb.add(self.tab2, text="Map Columns")
        self.nb.tab(self.tab2, state="disabled")
        
        ttk.Label(
            self.tab2,
            text="Created by Alex Oniciuc",
            anchor="e",
            background=	"#3E0E8C",
            foreground="white",
            font=(None, 8, "italic")
        ).place(relx=1.0, rely=1.0, anchor="se", x=-5, y=-5)
       
    

        # Tab 3: Results
        self.tab3 = ttk.Frame(self.nb)
        self.nb.add(self.tab3, text="Validation Results")
        self.nb.tab(self.tab3, state="disabled")
        self.result_container = ScrollableFrame(self.tab3)
        self.result_container.pack(fill="both", expand=True)

    def on_load(self):
        
          # ─── RESET UI PENTRU UN NOU LOAD ───
        self.sheet_frame.pack_forget()
        self.sheet_cb.config(values=[])
        self.sheet_var.set("")          # golește selecția anterioară
    # dezactivează tab‑urile 2 și 3
        self.nb.tab(self.tab2, state="disabled")
        self.nb.tab(self.tab3, state="disabled")
    # șterge conținutul din MappingTab (tab2)
        for w in self.tab2.winfo_children():
            w.destroy()
    # și din Results (tab3)
        for w in self.result_container.scrollable_frame.winfo_children():
            w.destroy()
        path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")]
    )
        if not path:
            return
        self.loaded_path = path

        from openpyxl import load_workbook
        try:
            wb = load_workbook(path, data_only=True)
        except (PermissionError, IOError):
            messagebox.showwarning(
                "File In Use",
                "The Excel file you selected is already open.\n"
                "Please close it before selecting it in the application."
            )
            return
        sheets = wb.sheetnames
        if len(sheets) > 1:
        # adaugă o opțiune goală la început, ca user‑ul să aleagă activ
            self.sheet_cb.config(values=[""] + sheets)
            self.sheet_frame.pack(pady=(0,20), fill="x")
        # NU mai facem self.sheet_var.set(sheets[0]) – lăsăm gol
            self.selected_sheet = None
            return


    # altfel (un singur sheet), ascunde picker-ul și continuă
        self.sheet_frame.pack_forget()
        self.selected_sheet = sheets[0]

    # încarcă datele și deschide Tab2
        self._load_dataframe()
        self.wb = wb

    # ... restul codului tău de inițializare a MappingTab și navigare ...
        expected = [
        "Country", "Brand", "Product ID", "Product Name (Local Language)",
        "Product Name (English)", "Product Description (Local Language)",
        "Product Description (English)", "Category", "Sub-Category",
        "Product Image URL", "Product URL", "MAP", "MSRP", "Product Video URL",
        ]
        for w in self.tab2.winfo_children():
            w.destroy()
        mt = MappingTab(self.tab2, self.cols, expected, self._on_validate)
        mt.pack(fill="both", expand=True)

          # ── AICI ── plasăm creditul după ce am curățat tab2
        ttk.Label(
            self.tab2,
            text="Created by Alex Oniciuc",
            anchor="e",
            background="#3E0E8C",
            foreground="white",
            font=(None, 8, "italic")
        ).place(relx=1.0, rely=1.0, anchor="se", x=-5, y=-5)

        
        self.nb.tab(self.tab2, state="normal")
        self.nb.select(self.tab2)

    def _on_sheet_selected(self, event):
    # 1) Setează sheet-ul nou
        self.selected_sheet = self.sheet_var.get()
    # 2) Încarcă datele (o singură citire) și normalizează ca în _load_dataframe
        import pandas as pd
        df = pd.read_excel(
            self.loaded_path,
            sheet_name=self.selected_sheet,
            dtype=str,
            header=0
        )
        df.columns = df.columns.str.strip()
        df.reset_index(drop=True, inplace=True)
        self.df = df
        self.cols = list(df.columns)
    # 3) Creează MappingTab și navighează la Tab2
    # (exact același cod pe care l-ai mutat din on_load)
        from openpyxl import load_workbook
        wb = load_workbook(self.loaded_path, data_only=True)
        self.wb = wb

        expected = [
        "Country","Brand","Product ID",  "Product Name (Local Language)",
        "Product Name (English)","Product Description (Local Language)",
        "Product Description (English)", "Category", "Sub-Category",
        "Product Image URL","Product URL","MAP", "MSRP","Product Video URL",
        ]
        for w in self.tab2.winfo_children():
            w.destroy()

        mt = MappingTab(self.tab2, self.cols, expected, self._on_validate)
        mt.pack(fill="both", expand=True)

        ttk.Label(
            self.tab2,
            text="Created by Alex Oniciuc",
            anchor="e",
            background="#3E0E8C",
            foreground="white",
            font=(None, 8, "italic")
            ).place(relx=1.0, rely=1.0, anchor="se", x=-5, y=-5)
       
        self.nb.tab(self.tab2, state="normal")
        self.nb.select(self.tab2)

    def _load_dataframe(self):
        import pandas as pd

        if not getattr(self, "loaded_path", None):
            return

        df = pd.read_excel(self.loaded_path, sheet_name=self.selected_sheet, dtype=str, header=0)
        df.columns = df.columns.str.strip()
        df.reset_index(drop=True, inplace=True)

        self.df   = df
        self.cols = list(df.columns)


   

    
    def _on_validate(self, to_rename, extra_id_cols, mapped_props):
        # print("DEBUG: _on_validate a fost apelat cu:", to_rename, extra_id_cols, mapped_props)

    # ── Reuse DataFrame și workbook încărcate anterior (fără re‑citire de pe disc)
        df = self.df.copy()
        wb = self.wb

    # ── Curățări și redenumiri
        if "Product Image URL" in df.columns and "Image" in df.columns:
            df = df.drop(columns=["Product Image URL"])
        df = df.rename(columns=to_rename)
        df = df.loc[:, ~df.columns.duplicated()]
        for prop in set(self.cols) | set(to_rename.values()):
            if prop not in df.columns:
                df[prop] = ""

    # ── Aici forțăm foaia selectată să devină „activă”
        sheet_index = wb.sheetnames.index(self.selected_sheet)
        wb.active = sheet_index

    # ── Validare cu fallback pe erori (apel neschimbat)
        try:
            report = validate_file(
                df_processed=df,
                wb_original=wb,
                extra_id_cols=extra_id_cols,
                mapped_props=mapped_props
            )
        except Exception as e:
            messagebox.showerror("Validation Error", f"A apărut o eroare neașteptată:\n{e}")
            return
        # import traceback; traceback.print_exc()

    # ── Fallback mapping dacă nu sunt suficiente mapping-uri
        if len(mapped_props) < 2:
            self.current_mapping = { col: col for col in df.columns }
        else:
            self.current_mapping = mapped_props

        self.current_report = report

    # ── Afișare rezultate
        self.display_results(report)


    def display_results(self, report_data):    
        # Șterge vechile rezultate
        for w in self.result_container.scrollable_frame.winfo_children():
            w.destroy()

         # ─── Afișăm numele fișierului, centrat deasupra File Format Checks ───
        from pathlib import Path
        if self.loaded_path:
            file_name = Path(self.loaded_path).name
            ttk.Label(
               self.result_container.scrollable_frame,
               text=file_name,
               font=("Arial", 14, "bold")
            ).pack(pady=(5,15), anchor="center")

        numeric = ["% Data Completeness", "% Data Uniqueness", "% Match Rate",
                   "Check Fail Count", "Empty Cell Count"]

        for group in report_data.get("validation_group_order", []):
            data = report_data.get(group, [])
            if not data:
                continue

            ttk.Label(
                self.result_container.scrollable_frame,
                text=group,
                font=("Arial",12,"bold")
            ).pack(anchor="w", pady=5)

            # Afișăm toate coloanele din raport
            cols = list(data[0].keys())
            # if "Check Fail Example Cell Reference" not in cols:
            #     cols.append("Check Fail Example Cell Reference")
            if group != "File Format Checks" and "Check Fail Example Cell Reference" not in cols:
                  cols.append("Check Fail Example Cell Reference")
            tree = ttk.Treeview(
                self.result_container.scrollable_frame,
                columns=cols,
                show="headings",
                height=min(len(data), 20)
               
            )

            
            for c in cols:
                tree.heading(c, text=c)
                tree.column(
                    c,
                    width=110 if c in numeric else 190,
                    anchor=("e" if c in numeric else "w")
                    
                )

            # ─── Tooltip doar pentru coloana "Check Fail Example" ───
            tip = ToolTip(tree)
            # cols_names = list(data[0].keys())
            columns = tree["columns"] 
           
            def on_motion(event):
                tv = event.widget   # ← folosim widget, nu tree
                if tv.identify("region", event.x, event.y) != "cell":
                    tip.hide()
                    return
                    
                # 2) Aflăm rândul & coloana
                rowid = tv.identify_row(event.y)
                colid = tv.identify_column(event.x)
                if not (rowid and colid):
                    tip.hide()
                    return

    # 3) Index + nume coloană
                idx = int(colid.lstrip("#")) - 1
                columns = tv["columns"]
                if idx < 0 or idx >= len(columns):
                    tip.hide()
                    return
                
                col_name = columns[idx]
                if "Check Fail Example" not in col_name:
                    tip.hide()
                    return
                
    # 4) Preluăm raw vals și afișăm tooltip-ul
                vals = tv.item(rowid, "values") or ()
                # print(f"[DEBUG] La hover {rowid}, raw_vals = {vals!r}")
                if idx >= len(vals):
                    tip.hide()
                    return
                text = vals[idx] 
                if text:
                    tip.show(text, event.x_root + 20, event.y_root + 10)
                else:
                    tip.hide()
  

            tree.bind("<Motion>", on_motion)
            tree.bind("<Leave>", lambda e: tip.hide())
# ───────────────────────────────────────────────────────

            for item in data:
                # Outcome și flags pentru ascundere
                outcome = item.get("Check Outcome", "")
                is_skip_not_mapped = (
                    outcome == "⏭️ Skip" and item.get("Explanation") == "Not mapped"
                )
                is_pass = outcome.startswith("✅")

                # Construim valorile, mascând celulele conform condițiilor:
                # - pentru skip+not mapped: ascundem toate coloanele în afara primelor 3
                # - pentru pass: ascundem doar Check Fail Count
                vals = []
                for c in cols:
                    if is_skip_not_mapped and c not in ("Check Performed", "Check Outcome", "Explanation"):
                        vals.append("")
                    #  elif is_pass and c == "Check Fail Count":
                    elif(
                            c == "Check Fail Count"
                            and (
                                outcome.startswith("✅")
                                or (outcome == "⏭️ Skip" and item.get("Explanation") == "No data present")
                                )
                            ):
                            vals.append("")

                
                        
                    else:
                        v = item.get(c, "")
                        if isinstance(v, (list, dict)):
                            vals.append(json.dumps(v, ensure_ascii=False))
                        else:
                            vals.append(v)

                iid = tree.insert("", "end", values=vals)
                
                # Aplicăm tag-uri pentru stil
                if outcome.startswith("✅"):
                    tree.item(iid, tags=("pass",))
                elif outcome.startswith("❌"):
                    tree.item(iid, tags=("fail",))
                else:
                    tree.item(iid, tags=("skip",))

            tree.tag_configure("pass", background="#d4edda")
            tree.tag_configure("fail", background="#f8d7da")
            tree.tag_configure("skip", background="#ececec")
            tree.pack(fill="both", expand=True, pady=(0,10))

       

            if group == "Data Format Checks":
                btn = ttk.Button(
                self.result_container.scrollable_frame,
                text="Download Fail Report",
                command=self._on_download_fail_report
                )
                btn.pack(pady=(0,15))

            self.nb.tab(self.tab3, state="normal")
            self.nb.select(self.tab3)

   


    def _on_download_fail_report(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files","*.xlsx")],
            title="Save Fail Report for Data Format Checks"
        )
        if not path:
            return
        
    # ——— Check if the source Excel file is locked ———
        try:
        # Try opening it for read+write; if it's open in Excel, this will fail
            with open(self.loaded_path, 'rb+'):
                pass
        except (PermissionError, IOError):
            messagebox.showwarning(
            "File Locked",
            f"The original Excel file:\n{self.loaded_path}\nis still open.\n"
            "Please close it before downloading the fail report."
            )
            return
         # reîncarc workbook-ul normal, nu read_only
        from openpyxl import load_workbook
        wb_for_export = load_workbook(self.loaded_path, data_only=True)

    # dacă ai mapat <2 coloane, fallback la toate coloanele din df
        if len(self.current_mapping) < 2:
        # creează un dict {col: col} pentru fiecare coloană
            mapping = { col: col for col in self.df.columns }
        else:
            mapping = self.current_mapping

        try:
            export_data_format_fails(
            report_data=self.current_report,
            df=self.df,
            wb_original=wb_for_export,
            mapping=mapping,
            save_path=path,
            sheet_name=self.selected_sheet
            )
            
            messagebox.showinfo("Export complete", f"Saved to:\n{path}")
        
       

# redeschidem raportul de fails tocmai salvat
            from openpyxl import load_workbook

# după ce ai salvat raportul și ai path 
            wb_exp = load_workbook(path, data_only=True)

# lista nodurilor de validare din raportul UI
            checks = ["Demo Data", "Special Characters", "Formulas", "HTML Tags"]

# citește fiecare sheet din fișierul de export
            for chk in checks:
    # sheet-urile din fișier au exact același nume
                if chk in wb_exp.sheetnames:
                    ws = wb_exp[chk]
        # calculează nr. de fails = nr. rânduri − 1 (header)
                    cnt = ws.max_row - 1
        # injectează în self.current_report
                    for item in self.current_report["Data Format Checks"]:
                        if item["Check Performed"] == chk:
                            item["Check Fail Count"] = cnt
                            break

# apoi afișează din nou
            self.display_results(self.current_report)
        except Exception as e:
            messagebox.showerror("Export error", str(e))
def main():
    app = OfflineCatalogValidatorApp()
    app.mainloop()

if __name__ == '__main__':
    main()

# if __name__ == '__main__':
#     app = OfflineCatalogValidatorApp()
#     app.mainloop()
