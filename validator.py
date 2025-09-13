import pandas as pd
import json
import re
import numpy as np

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from urllib.parse import urlparse


# import builtins

# DEBUG = False  # pune True când vrei să vezi din nou DEBUG‑urile

# # păstrăm referinţa la print-ul original
# _original_print = builtins.print

# if not DEBUG:
#     # suprascriem print în acest modul: devine no‑op
#     def print(*args, **kwargs):
#         pass
# else:
#     # readucem print-ul original
#     print = _original_print


def is_valid_url(url):
    if not url:
        return False
    parsed = urlparse(url)
    return parsed.scheme in ["http", "https"] and bool(parsed.netloc)

# --- Validation Logic (same as catalog_validator) ---

def trim_val(val):
    return val.strip() if isinstance(val, str) else val


def report_check(performed, outcome, **kwargs):
    result = {"Check Performed": performed, "Check Outcome": outcome}
    result.update(kwargs)
    return result

def group_a(wb, sheet_name=None):
    """
    File Format Checks – Single Worksheet & Hidden Rows/Cols
    – Single Worksheet: verifică dacă există exact o foaie.
    – Hidden Rows/Hidden Columns: se aplică doar pe sheet_name (dacă există în wb),
      altfel pe wb.active.
    – Nu se mai raportează “Check Fail Example Cell Reference”.
    """
    results = []

    # --- 1) Single Worksheet (global) ---
    sheets = wb.sheetnames
    outcome_sw = '✅ Pass' if len(sheets) == 1 else '❌ Fail'
    results.append(report_check(
        "Single Worksheet",
        outcome_sw,
        Explanation=f"{len(sheets)} sheets: {sheets}"
    ))

    # --- 2) Alegem foaia pe care lucrăm pentru hidden rows/cols ---
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    title = ws.title

    # --- 3) Hidden Rows ---
    hidden_rows = [i for i, dim in ws.row_dimensions.items() if dim.hidden]
    outcome_hr = '❌ Fail' if hidden_rows else '✅ Pass'
    # doar Count și Example (fără Cell Reference)
    hr_kwargs = {
        "Check Fail Count": len(hidden_rows),
        "Check Fail Example": str(hidden_rows[0]) if hidden_rows else ""
    }
    results.append(report_check(
        "Hidden Rows",
        outcome_hr,
        Explanation=(f"[{title}] Hidden rows: {hidden_rows}" if hidden_rows else ""),
        **hr_kwargs
    ))

    # --- 4) Hidden Columns ---
    hidden_cols = [c for c, dim in ws.column_dimensions.items() if dim.hidden]
    outcome_hc = '❌ Fail' if hidden_cols else '✅ Pass'
    hc_kwargs = {
        "Check Fail Count": len(hidden_cols),
        "Check Fail Example": hidden_cols[0] if hidden_cols else ""
    }

    results.append(report_check(
        "Hidden Columns",
        outcome_hc,
        Explanation=(f"[{title}] Hidden cols: {hidden_cols}" if hidden_cols else ""),
        **hc_kwargs
    ))

    return results



from openpyxl.utils import get_column_letter
import re


def group_b(df, wb, mapped_props=None):
    """
    Data Format Checks (Demo Data, Special Characters, Formulas, HTML Tags)
    – rulează **numai** pe foaia activă (wb.active), nu pe toate foile.
    – fiecare fail e raportat sub forma SheetName!ColLetterRow.
    (Optimizare fără schimbare de funcționalitate sau UI:
     - evită recăutarea literei coloanei pentru fiecare regulă
     - accesează valorile cu iter_rows(values_only=True) în loc de indexări celulă‑cu‑celulă)
    """
    checks = [
        ("Demo Data",        re.compile(r'\bdemo(?:brand|sku|_category)?\b', re.IGNORECASE)),
        ("Special Characters", set("©$€£¥™®@")),
        ("Formulas",         "="),
        ("HTML Tags",        re.compile(r'<[^>]+>|&lt;[^&]+&gt;'))
    ]
    results = []

    ws = wb.active
    sheet_name = ws.title

    # Dacă nu avem coloane în DataFrame, păstrăm exact același comportament
    if df.columns.empty:
        for name, _ in checks:
            results.append(report_check(
                name, "⏭️ Skip", Explanation="No columns in DataFrame",
                **{"Check Fail Count": 0,
                   "Check Fail Example": "",
                   "Check Fail Example Cell Reference": ""}
            ))
        return results

    # mapăm header-ele din rândul 1 -> literă de coloană
    header_to_letter = {}
    for cell in ws[1]:
        header = str(cell.value).strip() if cell.value is not None else ""
        if header in df.columns:
            header_to_letter[header] = cell.column_letter

    # PRE-SCAN HTML (ordinea și logica rămân identice)
    html_pattern = re.compile(r'<[^>]+>|&lt;[^&]+&gt;')
    html_fails = []
    for header, col_letter in header_to_letter.items():
        # obținem toată coloana ca valori simple, în ordinea rândurilor
        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(col_letter)
        vals_iter = ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True)
        row_num = 2
        for (raw,) in vals_iter:
            text = str(raw or "").strip()
            if text:
                for tag in html_pattern.findall(text):
                    html_fails.append((sheet_name, col_letter, row_num, tag))
            row_num += 1

    # Bucla inițială per regulă -> per coloană -> per rând (păstrăm aceeași ordine)
    for name, pattern in checks:
        fails = list(html_fails) if name == "HTML Tags" else []

        for header in df.columns:
            col_letter = header_to_letter.get(header)
            if not col_letter:
                continue

            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(col_letter)
            vals_iter = ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True)

            row_num = 2
            for (raw,) in vals_iter:
                text = str(raw or "").strip()
                if not text:
                    row_num += 1
                    continue

                if name == "Demo Data":
                    for m in pattern.finditer(text):
                        fails.append((sheet_name, col_letter, row_num, m.group(0)))

                elif name == "Special Characters":
                    for ch in pattern:
                        for _ in range(text.count(ch)):
                            fails.append((sheet_name, col_letter, row_num, ch))

                elif name == "Formulas":
                    # păstrăm EXACT logica inițială bazată pe text
                    if text.startswith("="):
                        fails.append((sheet_name, col_letter, row_num, text))

                row_num += 1

        outcome = "❌ Fail" if fails else "✅ Pass"
        if fails:
            _, col, row, snippet = fails[0]
            example_ref = f"{sheet_name}!{col}{row}"
        else:
            snippet = example_ref = ""

        results.append(report_check(
            name,
            outcome,
            Explanation="",
            **{
                "Check Fail Count": len(fails),
                "Check Fail Example": snippet,
                "Check Fail Example Cell Reference": example_ref
            }
        ))

    return results
      # mapăm fiecare header din df la litera de coloană Excel
    header_to_letter = {}
    for cell in ws[1]:
        header = str(cell.value).strip() if cell.value is not None else ""
        if header in df.columns:
            header_to_letter[header] = cell.column_letter
      # ——— PRE-SCAN PENTRU HTML TAGS ———
       # ——— PRE-SCAN PENTRU HTML TAGS (doar pe coloanele din df) ———
    html_pattern = re.compile(r'<[^>]+>|&lt;[^&]+&gt;')
    html_fails = []
    for header, col_letter in header_to_letter.items():
        for row in range(2, ws.max_row + 1):
            raw = ws[f"{col_letter}{row}"].value
            text = str(raw or "").strip()
            if not text:
                continue
            for tag in html_pattern.findall(text):
                html_fails.append((sheet_name, col_letter, row, tag))
    # ————————————————————————————————————————————————
    # Pentru fiecare regulă, scanăm coloanele din df pe foaia activă
    for name, pattern in checks:
        # fails = []
        if name == "HTML Tags":
            fails = list(html_fails)
        else:
            fails = []
        
        for header in df.columns:
            # găsim litera coloanei în Excel după header
            col_letter = None
            for cell in ws[1]:
                if cell.value and str(cell.value).strip() == str(header).strip():
                    col_letter = cell.column_letter
                    break
            if not col_letter:
                continue

            # iterăm rândurile din acea coloană
            for row in range(2, ws.max_row + 1):
                raw = ws[f"{col_letter}{row}"].value
                text = str(raw or "").strip()
                if not text:
                    continue

                if name == "Demo Data":
                    for m in pattern.finditer(text):
                        fails.append((sheet_name, col_letter, row, m.group(0)))

                elif name == "Special Characters":
                    for ch in pattern:
                        for _ in range(text.count(ch)):
                            fails.append((sheet_name, col_letter, row, ch))

                elif name == "Formulas":
                    if text.startswith("="):
                        fails.append((sheet_name, col_letter, row, text))

                # elif name == "HTML Tags":
                #     for tag in pattern.findall(text):
                #         fails.append((sheet_name, col_letter, row, tag))
                

        # Construim raportul
        outcome = "❌ Fail" if fails else "✅ Pass"
        if fails:
            _, col, row, snippet = fails[0]
            example_ref = f"{sheet_name}!{col}{row}"
        else:
            snippet = example_ref = ""

        results.append(report_check(
            name,
            outcome,
            # Explanation=(f"{len(fails)} issue(s) found" if fails else "")
            Explanation="",
            **{
                "Check Fail Count": len(fails),
                "Check Fail Example": snippet,
                "Check Fail Example Cell Reference": example_ref
            }
        ))

    return results



from openpyxl.utils import get_column_letter
import pandas as pd


def completeness_check(df, col):
    total = len(df)

    # 1) Obținem datele
    raw = df[col].fillna("").map(trim_val)
    mask = raw.astype(bool)
    if isinstance(mask, pd.DataFrame):
        mask = mask.any(axis=1)

    # 2) Calculăm pct
    non_empty = int(mask.sum())
    pct       = f"{int(non_empty/total*100)}%" if total else "0%"

    # 3) Câte lipsesc
    fails = total - non_empty

    example, ref = "", ""
    if fails > 0:
        first_blank_idx = mask[~mask].index[0]
        cell_val = df.at[first_blank_idx, col]

        # ─── Tratem Series separat ────────────────────────────────────────
        if isinstance(cell_val, pd.Series):
            # aplicăm trim_val pe fiecare coloană duplicată,
            # apoi luăm primul non-blank
            trimmed = (
                cell_val
                .map(lambda x: trim_val(x) if isinstance(x, str) else "")
                .loc[lambda s: s != ""]
            )
            example = trimmed.iloc[0] if not trimmed.empty else "<blank>"

        else:
            # ─── Acum e scalar ───────────────────────────────────────────────
            if pd.isna(cell_val):
                example = "<blank>"
            elif isinstance(cell_val, str):
                t = trim_val(cell_val)
                example = "<blank>" if t == "" else t
            else:
                example = str(cell_val)

        # ─── Construim referința A1 ───────────────────────────────────────
        col_idx    = list(df.columns).index(col) + 1
        col_letter = get_column_letter(col_idx)
        ref        = f"{col_letter}{first_blank_idx+2}"

    return report_check(
        f"{col} Completeness",
        "✅ Pass" if fails == 0 else "❌ Fail",
        **{
            "% Data Completeness": pct,
            "Explanation": "" if fails == 0 else "Missing",
            "Check Fail Count": fails,
            "Check Fail Example": example,
            "Check Fail Example Cell Reference": ref
        }
    )



def uniqueness_check(df, col):
    vals = df[col].fillna("").map(trim_val)
    non_empty = vals[vals != ""]

    dup = non_empty[non_empty.duplicated(keep=False)]
    total = len(non_empty)
    fails = len(dup)
    pct = f"{int((total - fails) / total * 100)}%" if total else "100%"

    example, ref = "", ""
    if fails:
        idx = dup.index[0]
        example = df.at[idx, col]
        # afișăm referința celulei folosind get_column_letter
        col_idx    = list(df.columns).index(col) + 1
        col_letter = get_column_letter(col_idx)
        ref        = f"{col_letter}{idx+2}"

    return report_check(
        f"{col} Uniqueness",
        '✅ Pass' if fails == 0 else '❌ Fail',
        **{
            "% Data Uniqueness": pct,
            "Explanation": "" if fails == 0 else "Duplicates",
            "Check Fail Count": fails,
            "Check Fail Example": example,
            "Check Fail Example Cell Reference": ref
        }
    )


def group_c(df, mapped_props=None):
    """
    Mandatory Data - Completeness Checks
    Verifică completitudinea coloanelor obligatorii doar dacă au fost mapate.
    """
    mandatory = [
        "Country",
        "Brand",
        "Product ID",
        "Product Name (English)",
        "Product Image URL"
    ]
    results = []
    mapped = set(mapped_props or [])

    for col in mandatory:
        check_name = f"{col} Completeness"

        # 1) Dacă mapped_props există și câmpul nu e printre ele → Skip
        if mapped and col not in mapped:
            results.append(report_check(
                check_name,
                "⏭️ Skip",
                Explanation="Not mapped",
                **{"% Data Completeness": "0%",
                   "Check Fail Count": 0,
                   "Check Fail Example": "",
                   "Check Fail Example Cell Reference": ""}
            ))

        # 2) Dacă e mapat, dar nu există coloana în df → Skip
        elif mapped and col not in df.columns:
            results.append(report_check(
                check_name,
                "⏭️ Skip",
                Explanation="Mapped but column missing",
                **{"% Data Completeness": "0%",
                   "Check Fail Count": 0,
                   "Check Fail Example": "",
                   "Check Fail Example Cell Reference": ""}
            ))

        # 3) Altfel, aplicăm completeness_check
        else:
            results.append(completeness_check(df, col))

    return results



def group_d(df, mapped_props=None):
    """
    Mandatory Data - Uniqueness Checks
    Verifică unicitatea pe câmpurile obligatorii doar dacă au fost mapate.
    Dacă nu sunt mapate: ⏭️ Skip (Not mapped)
    Dacă sunt mapate, dar coloana nu există în df: ⏭️ Skip (Mapped but column missing)
    Altfel: uniqueness_check.
    """
    mandatory = [
        "Product ID",
        "Product Name (English)"
    ]
    results = []
    mapped = set(mapped_props or [])

    for col in mandatory:
        check_name = f"{col} Uniqueness"
        # 1) dacă există mapped_props dar col nu e în ele → Skip
        if mapped and col not in mapped:
            results.append(report_check(
                check_name,
                "⏭️ Skip",
                Explanation="Not mapped",
                **{
                    "% Data Uniqueness": "",
                    "Check Fail Count": 0,
                    "Check Fail Example": "",
                    "Check Fail Example Cell Reference": ""
                }
            ))
        # 2) dacă col e în mapped_props dar nu există în df → Skip
        elif mapped and col not in df.columns:
            results.append(report_check(
                check_name,
                "⏭️ Skip",
                Explanation="Mapped but column missing",
                **{
                    "% Data Uniqueness": "",
                    "Check Fail Count": 0,
                    "Check Fail Example": "",
                    "Check Fail Example Cell Reference": ""
                }
            ))
        # 3) Altfel, rulăm verificarea de unicitate
        else:
            results.append(uniqueness_check(df, col))

    return results



import re
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


def group_e(df, mapped_props=None):
    """
    Mandatory Data - Country Unique Count
    Verifică că există exact o singură valoare nenulă unică în coloana Country,
    doar dacă Country a fost mapat. Altfel: ⏭️ Skip.
    """
    col = "Country"
    check_name = "Country Unique Count"
    mapped = set(mapped_props or [])

    # 1) Dacă există mapped_props dar Country nu e printre ele → Skip
    if mapped and col not in mapped:
        return [report_check(
            check_name,
            "⏭️ Skip",
            Explanation="Not mapped",
            **{
                "Count Unique": 0,
                "List Unique Values": "",
                "Check Fail Example": "",
                "Check Fail Example Cell Reference": ""
            }
        )]

    # 2) Dacă s-a mapat, dar coloana lipsește din df → Skip
    if mapped and col not in df.columns:
        return [report_check(
            check_name,
            "⏭️ Skip",
            Explanation="Mapped but column missing",
            **{
                "Count Unique": 0,
                "List Unique Values": [],
                "Check Fail Example": "",
                "Check Fail Example Cell Reference": ""
            }
        )]
    
    # ─── 2.1) FAIL dacă avem rânduri blank în Country ───
    blank_mask = df[col].isna() | df[col].astype(str).map(trim_val).eq("")
    if blank_mask.any():
    # prima linie blank
        first_idx = df.index[blank_mask][0]
        excel_row = first_idx + 2
        col_letter = get_column_letter(df.columns.get_loc(col) + 1)
        return [report_check(
            check_name,
            "❌ Fail",
            Explanation="Missing",
            **{
            "Count Unique": "",
            "List Unique Values": "",
            "Check Fail Example": "<blank>",
            "Check Fail Example Cell Reference": f"{col_letter}{excel_row}"
        }
    )]
    #     return [report_check(
    #         check_name,
    #         "❌ Fail",
    #         Explanation="Blank country entries found",
    #         **{
    #             "Count Unique": 0,
    #             "List Unique Values": "",
    #             "Check Fail Example": "",
    #             "Check Fail Example Cell Reference": f"{col_letter}{excel_row}"
    #     }
    # )]


    # 3) Altfel, aplicăm logica originală pe df (deja mapat)
    common = {
        "Count Unique": 0,
        "List Unique Values": [],
        "Check Fail Example": "",
        "Check Fail Example Cell Reference": ""
    }

    raw = df[col].dropna().map(str).map(trim_val).loc[lambda s: s != ""]
    parts = (
        raw.str.split(r"\+")
           .explode()
           .map(trim_val)
           .loc[lambda s: s != ""]
    )
    vals = parts.unique().tolist()
    # If Country column has no data, skip this check
    if not vals:
        return [report_check(
            check_name,
            "⏭️ Skip",
            Explanation="No data present",
            **{
                "Count Unique": 0,
                "List Unique Values": "",
                "Check Fail Example": "",
                "Check Fail Example Cell Reference": ""
            }
        )]
    list_display = ", ".join(vals) if vals else ""
    common["Count Unique"] = len(vals)
    common["List Unique Values"] = list_display

    if len(vals) == 1:
        outcome, explanation = "✅ Pass", ""
    else:
        outcome, explanation = "❌ Fail", f"Found {len(vals)} country codes"
        example = vals[0]
        common["Check Fail Example"] = example
        mask = raw.str.contains(fr"(?:^|(?<=\+))\s*{re.escape(example)}\s*(?=(?:\+|$))",regex=True)
        first_idx = df.index[mask][0]
        excel_row = first_idx + 2
        col_letter = get_column_letter(df.columns.get_loc(col) + 1)
        common["Check Fail Example Cell Reference"] = f"{col_letter}{excel_row}"

    row = report_check(
        check_name,
        outcome,
        Explanation=explanation,
        **common
    )
    return [row]


from urllib.parse import urlparse, urlunparse, unquote
from openpyxl.utils import get_column_letter

def normalize_url(raw: str) -> str:
    """
    Decodifică percent-encoding, strip-uiește whitespace,
    și pune scheme și host în lowercase.
    """
    if not raw:
        return ""
    p = urlparse(raw.strip())
    path  = unquote(p.path)
    query = unquote(p.query)
    return urlunparse((
        p.scheme.lower(),
        p.netloc.lower(),
        path,
        p.params,
        query,
        ""
    ))

from urllib.parse import urlparse
from openpyxl.utils import get_column_letter
import numpy as np
import pandas as pd
import re

VALID_URL_REGEX = re.compile(r'^(https?://)[A-Za-z0-9\.-]+\.[A-Za-z]{2,}.*$')

def group_f(wb, df, mapped_props=None):
    """
    Mandatory Data - URL Field Checks for Product Image URL,
    cu fallback pe hyperlink-ul atașat în Excel dacă textul nu e URL.
    """
    col_name   = "Product Image URL"
    check_name = f"{col_name} Hyperlink Check"
    mapped     = set(mapped_props or [])

    # A) Skip dacă n-a fost mapată
    if mapped and col_name not in mapped:
        return [report_check(
            check_name, "⏭️ Skip", Explanation="Not mapped",
            **{"% Pass Rate":"0%","Check Fail Count":0,"Check Fail Example":"","Check Fail Example Cell Reference": ""}
        )]
    # B) Skip dacă e mapată dar lipsește coloana
    if mapped and col_name not in df.columns:
        return [report_check(
            check_name, "⏭️ Skip", Explanation="Mapped but column missing",
            **{"% Pass Rate":"0%","Check Fail Count":0,"Check Fail Example":"","Check Fail Example Cell Reference": ""}
        )]

    # C) Extragem și curățăm valorile
    raw = df[col_name]
    if isinstance(raw, pd.DataFrame):
        raw = raw.iloc[:,0]
    vals = raw.fillna("").astype(str).str.strip()
    # eliminăm echo-ul header-ului și duplicatele
    vals = vals.loc[(vals != col_name) & ~vals.duplicated()]

    # D) Fail “Missing” dacă nu rămâne niciun URL
    if vals.eq("").all():
        idx0 = df.index[0] + 2
        col_idx   = list(df.columns).index(col_name) + 1
        col_letter= get_column_letter(col_idx)
        return [report_check(
            check_name, "❌ Fail", Explanation="Missing",
            **{
              "% Pass Rate":"0%",
              "Check Fail Count": len(df),
              "Check Fail Example": "<blank>",
              "Check Fail Example Cell Reference": f"{col_letter}{idx0}"
            }
        )]

    # E) Validare combinată: text‑based + fallback hyperlink
    ws = wb.active
       # alegem foaia pe care a selectat-o user-ul (fallback la primul dacă nu există)
   
    # map ref → hyperlink.target
    hmap = {hl.ref: hl.target.strip() if hl.target else "" for hl in ws._hyperlinks}

    # litera coloanei
    col_idx    = list(df.columns).index(col_name) + 1
    col_letter = get_column_letter(col_idx)

    # 1) Text‑based URL check (regex strict)
    mask_text_valid = vals.str.match(VALID_URL_REGEX).to_numpy()

    # 2) Hyperlink‑based check
    coords = [f"{col_letter}{i+2}" for i in vals.index]
    links_series   = pd.Series({c: hmap.get(c, "") for c in coords})
    mask_link_valid = links_series.str.match(VALID_URL_REGEX).to_numpy()

    # 3) Unificăm: valid dacă
    mask_valid = mask_text_valid | mask_link_valid

    total     = mask_valid.size
    match_cnt = int(mask_valid.sum())
    fails_idx = np.where(~mask_valid)[0]

    pct     = f"{int(match_cnt/total*100)}%" if total else "0%"
    outcome = "✅ Pass" if fails_idx.size == 0 else "❌ Fail"
    details = {"% Pass Rate": pct, "Check Fail Count": len(fails_idx)}

    if fails_idx.size:
        i       = fails_idx[0]
        row     = vals.index[i] + 2
        val     = vals.iat[i]
        link    = links_series.iat[i]
        col_ref = f"{col_letter}{row}"

        details["Check Fail Example"]               = "<blank>" if val == "" else val
        details["Check Fail Example Cell Reference"] = col_ref

        # motiv detaliat în Explanation
        if mask_text_valid[i]:
            # text e ok → hyperlink invalid
            explanation = f"Invalid hyperlink target ({link}) at {col_ref}"
        elif link:
            # text invalid, dar hyperlink atașat e ok
            explanation = f"Text not URL but hyperlink used at {col_ref}"
        else:
            # nici text, nici hyperlink
            explanation = f"No hyperlink attached at {col_ref}"
    else:
        explanation = "All cells have a valid URL or hyperlink target."

    return [report_check(
        check_name,
        outcome,
        Explanation=explanation,
        **details
    )]


import math
def sanitize(obj):
    """
    Recursively sanitize the output:
    - Convert numpy integers to Python ints
    - Convert numpy floats to Python floats, mapping NaN to None
    - Convert NaN floats to None
    """
    if isinstance(obj, dict):
        return {k: sanitize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [sanitize(v) for v in obj]
    # numpy integer
    if isinstance(obj, np.integer):
        return int(obj)
    # numpy floating or Python float
    if isinstance(obj, (np.floating, float)):
        val = float(obj)
        if math.isnan(val):
            return None
        return val
    return obj


from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd

def validate_file(df_processed, wb_original, extra_id_cols=None, mapped_props=None):
    """
    df_processed   : pandas.DataFrame citit și redenumit conform mapping-ului
    wb_original    : openpyxl.Workbook încărcat din .xlsx-ul original
    extra_id_cols  : listă de proprietăți suplimentare pentru identificatori
    mapped_props   : dict {proprietate_canonicală: nume_coloană_în_df}

    """

    if mapped_props is None:
        mapped_props = {}
    if extra_id_cols is None:
        extra_id_cols = []

    # 1) Workbook și DataFrame preluate din argumente
    wb = wb_original
    df = df_processed.copy()

    # 2) Normalizează și redenumește header-ele
    df.columns = df.columns.str.strip().str.replace(r'\s+', ' ', regex=True)
    df.rename(columns={
        'Product  ID': 'Product ID',
        'Product ID ': 'Product ID',
        'Product Name(Local Language)': 'Product Name (Local)',
        'Product Name(English)': 'Product Name (English)',
        'Image': 'Product Image URL',
    }, inplace=True)

    df_processed = df_processed.loc[:, ~df_processed.columns.duplicated()]

    # 3) Rulează grupurile de validări
    output = {
        "validation_group_order": [
            "File Format Checks",
            "Data Format Checks",
            "Mandatory Data - Completeness Checks",
            "Mandatory Data - Uniqueness Checks",
            "Mandatory Data - Country Uniqueness Checks",
            "Mandatory Data - URL Field Checks",
            "Optional Data - Completeness Checks",
            "Optional Data - Uniqueness Checks",
            "Optional Data - URL Field Checks",
            "Product Name English - Mandatory Field - Character Limit Check",
            "Product Name Local - Optional Field - Character Limit Check",
            "Product Descriptions - Optional Fields - Character Limit Check",
            "Mandatory Data - Single ProductID Per Cell",
            "Optional Data - Single Secondary Product Identifier Per Cell",
            "Optional Data - Category Length & Tag Character Checks"
        ],
        "file_summary": {"row_count": len(df)},

        # 1) Verificări pe workbook complet
        "File Format Checks":                          group_a(wb),

        # 2) Data‐format pe df + mapped_props
        "Data Format Checks":                          group_b(df, wb, mapped_props=None),

        # 3) Mandatory completeness & uniqueness
        "Mandatory Data - Completeness Checks":        group_c(df, mapped_props=mapped_props),
        "Mandatory Data - Uniqueness Checks":          group_d(df, mapped_props=mapped_props),

        # 4) Max Allowable (country unique)
        "Mandatory Data - Country Uniqueness Checks":       group_e(df, mapped_props=mapped_props),

        # 5) URL Field Checks (imagine)
        "Mandatory Data - URL Field Checks":           group_f(wb, df, mapped_props=mapped_props),

        # 6) Optional data (completeness, uniqueness, URL)
        "Optional Data - Completeness Checks":         group_g(df,extra_id_cols=extra_id_cols, mapped_props=mapped_props),
        "Optional Data - Uniqueness Checks":           group_h(df,extra_id_cols=extra_id_cols, mapped_props=mapped_props),
        "Optional Data - URL Field Checks":            group_i(wb, mapped_props=mapped_props),

        # 7) Character‐limit checks
        "Product Name English - Mandatory Field - Character Limit Check":
                                                      group_j(df, mapped_props=mapped_props),
        "Product Name Local - Optional Field - Character Limit Check":
                                                      group_k(df, mapped_props=mapped_props),
        "Product Descriptions - Optional Fields - Character Limit Check":
                                                      group_l(df, mapped_props=mapped_props),

        # 8) Single‐ID checks
        "Mandatory Data - Single ProductID Per Cell":  group_m(df, mapped_props=mapped_props),
        "Optional Data - Single Secondary Product Identifier Per Cell":
                                                      group_n(df, extra_id_cols, mapped_props=mapped_props),

        # 9) Category data
        "Optional Data - Category Length & Tag Character Checks":        group_o(df, mapped_props=mapped_props)
    }

    return sanitize(output)



def completeness_with_locations(df, col):
    """
    Variantează completeness_check pentru a include și toate referințele
    celulelor goale (ex: G2, G5, ...).
    """
    total = len(df)
    # 1) Gasim care sunt goale (după strip)
    trim = df[col].fillna("").map(trim_val)
    blank_mask = trim == ""
    blank_idxs = blank_mask[blank_mask].index.tolist()  # index zero-based

    # 2) Calcul procent
    non_empty = total - len(blank_idxs)
    pct = int(non_empty / total * 100) if total else 100

    # 3) Construim lista de referinte Excel
    col_letter = chr(65 + df.columns.get_loc(col))
    refs = [f"{col_letter}{i+2}" for i in blank_idxs]

    # 4) Pregatim dict-ul
    result = {
        "Check Performed": f"{col} Completeness",
        "% Data Completeness": f"{pct}%",
    }
    if blank_idxs:
         # grab first blank cell
        first_idx = blank_idxs[0]
        first_ref = refs[0]
        raw_example = df.at[first_idx, col]
        # Dacă e NaN, None sau șir gol/spații → "<blank>"
        if pd.isna(raw_example) or raw_example is None or str(raw_example).strip() == "":
            example_val = "<blank>"
        else:
            example_val = raw_example
            # Explanation singular/plural

        count = len(blank_idxs)
        if count == 1:
            result["Explanation"] = "1 empty cell"
        else:
            result["Explanation"] = f"{count} empty cells"

        result.update({
            "Check Outcome": "❌ Fail",
            "Empty Cell Count": len(blank_idxs),
            "Empty Cell References": ", ".join(refs),
            "Check Fail Example": example_val,
            "Check Fail Example Cell Reference": first_ref
        })
    else:
        result["Explanation"] = "All cells populated"
        result.update({
            "Check Outcome": "✅ Pass",
            "Empty Cell Count": 0,
            "Empty Cell References": "",
            "Check Fail Example": "",
            "Check Fail Example Cell Reference": ""
        })

    return result

def group_g(df, mapped_props=None, extra_id_cols=None):
    """
    Optional Data - Completeness Checks
    Verifică completitudinea coloanelor opționale (inclusiv “Other” sau orice cod suplimentar)
    și raportează pentru fiecare atât % completeness,
    cât și numărul și locațiile celulelor goale.
    """
    import pandas as _pd

    mapped_props  = mapped_props or {}
    extra_id_cols = extra_id_cols or []

    # 1) Lista standard de proprietăți opționale
    optional_cols = [
        "SKU", "EAN", "UPC", "GTIN", "CTIN", "ASIN",
        "Product Name (Local Language)",
        "Product Description (English)",
        "Product Description (Local Language)",
        "Category", "Sub-Category",
        "Product URL", "Product Video URL",
        "MAP", "MSRP"
    ]

    # 2) Construim lista finală de prop_name:
    #    a) coloanele standard care au fost mapate
    #    b) plus orice cod suplimentar (din extra_id_cols), inclusiv “Other”
    all_props = [p for p in optional_cols   if p in mapped_props] + \
                [p for p in extra_id_cols    if p not in optional_cols]

    results = []
    for prop in all_props:
        check_name = f"{prop} Completeness"
        # define common înainte de orice continue, ca să nu dea eroare
        common = {
            "% Data Completeness":          "0%",
            "Check Fail Count":             0,
            "Check Fail Example":           "",
            "Check Fail Example Cell Reference": ""
        }
        # --- DEBUG (opțional) ---
        # print(f"DEBUG group_g: prop={prop!r}, mapped_props={mapped_props}, extra_id_cols={extra_id_cols}")

        # a) Not mapped → Skip
        if prop not in mapped_props:
            results.append({
                "Check Performed": check_name,
                "Check Outcome":    "⏭️ Skip",
                "% Data Completeness": common["% Data Completeness"],
                "Explanation":       "Not mapped",
                **common
            })
            continue

        # b) Column missing → Skip
        col = prop
        if col not in df.columns:
            results.append({
                "Check Performed": check_name,
                "Check Outcome":    "⏭️ Skip",
                "% Data Completeness": common["% Data Completeness"],
                "Explanation":       "Column missing",
                **common
            })
            continue

        # c) No data present → Skip
        vals = df[col].fillna("").astype(str).map(trim_val)
        if vals.eq("").all():
            results.append({
                "Check Performed": check_name,
                "Check Outcome":    "⏭️ Skip",
                "% Data Completeness": common["% Data Completeness"],
                "Explanation":       "No data present",
                **common
            })
            continue

        # d) Calculăm % completeness și count empty
        total       = len(vals)
        empty_count = int((vals == "").sum())
        pct_value   = int((total - empty_count) / total * 100) if total else 0

        common["% Data Completeness"] = f"{pct_value}%"
        common["Check Fail Count"]    = empty_count

        if empty_count > 0:
            # primul index gol
            first_idx   = vals[vals == ""].index[0]
            letter      = get_column_letter(df.columns.get_loc(col) + 1)
            cell_ref    = f"{letter}{first_idx + 2}"
            example_val = df.at[first_idx, col]
            # afișăm <blank> dacă valoarea e doar spații sau NaN
            example_display = (
                "<blank>"
                if (_pd.isna(example_val) or str(example_val).strip() == "")
                else str(example_val)
            )
            common["Check Fail Example"]                 = example_display
            common["Check Fail Example Cell Reference"] = cell_ref

            outcome, explanation = (
                "❌ Fail",
                f"{empty_count} empty cell" + ("s" if empty_count > 1 else "")
            )
        else:
            outcome, explanation = "✅ Pass", "All cells populated"

        results.append({
            "Check Performed":     check_name,
            "Check Outcome":        outcome,
            "% Data Completeness":  common["% Data Completeness"],
            "Explanation":          explanation,
            **common
        })

    return results




def group_h(df, mapped_props=None, extra_id_cols=None):
    """
    Optional Data - Uniqueness Checks
    → Verifică duplicatele DOAR pentru coloanele opționale mapate,
      plus orice prop suplimentar (înclusiv 'Other').
    """
    mapped_props  = mapped_props or {}
    extra_id_cols = extra_id_cols or []
    df = df.loc[:, ~df.columns.duplicated()]

    # 1) Proprietăți opționale standard
    base_props = [
        "SKU", "EAN", "UPC", "GTIN", "CTIN", "ASIN",
        "Product Name (Local Language)",
        "Product Description (English)",
        "Product Description (Local Language)",
        "Category", "Sub-Category",
        "Product URL", "Product Video URL",
        "MSRP", "MAP", "Product Image URL"
    ]

    # 2) Adăugăm în coadă orice cod suplimentar mapat (ex. coloana pentru “Other”)
    all_props = base_props + [p for p in extra_id_cols if p not in base_props]

    results = []
    for prop in all_props:
        check_name = f"{prop} Uniqueness"
        common = {
            "% Data Uniqueness": "",
            "Check Fail Count": 0,
            "Check Fail Example": "",
            "Check Fail Example Cell Reference": ""
        }

        # 3) Not mapped → Skip
        if prop not in mapped_props:
            results.append(report_check(
                check_name, "⏭️ Skip",
                Explanation="Not mapped",
                **common
            ))
            continue

        # 4) Mapped but header missing → Skip
        col = prop
        if col not in df.columns:
            results.append(report_check(
                check_name, "⏭️ Skip",
                Explanation="Column missing",
                **common
            ))
            continue

        # 5) Mapped & present but all values blank → Skip
        vals = df[col].fillna("").astype(str).map(trim_val)
        if vals.eq("").all().all():
            results.append(report_check(
                check_name, "⏭️ Skip",
                Explanation="No data present",
                **common
            ))
            continue

        # 6) Excepție Category/Sub-Category → întotdeauna Pass
        if prop in ("Category", "Sub-Category"):
            results.append(report_check(
                check_name,
                "✅ Pass",
                Explanation="Duplicates allowed for categories",
                **{
                    "% Data Uniqueness": "100%",
                    "Check Fail Count": 0,
                    "Check Fail Example": "",
                    "Check Fail Example Cell Reference": ""
                }
            ))
            continue

        # 7) Verificare de duplicate non-empty
        dupes = vals[vals.duplicated(keep=False) & vals.ne("")]
        pct = int((1 - dupes.nunique() / len(vals)) * 100) if len(vals) else 0
        
        # common["% Data Uniqueness"] = f"{100 * (1 - dupes.nunique() / len(vals)):.0f}%"
        common["Check Fail Count"] = len(dupes)

        if not dupes.empty:
            # luăm primul exemplar ca exemplu
            first_idx     = dupes.index[0] + 2  # +2 pt header + zero-based
            example_value = vals.iloc[dupes.index[0]]
            letter        = get_column_letter(df.columns.get_loc(col) + 1)
            common["Check Fail Example"]                 = example_value
            common["Check Fail Example Cell Reference"] = f"{letter}{first_idx}"
            outcome, expl = "❌ Fail", "Duplicate values found"
        else:
            outcome, expl = "✅ Pass", ""

        results.append(report_check(
            check_name, outcome,
            Explanation=expl,
            **common
        ))

    return results


from urllib.parse import urlparse
from openpyxl.utils import get_column_letter
import re

VALID_URL_REGEX = re.compile(r'^(https?://)[A-Za-z0-9\.-]+\.[A-Za-z]{2,}.*$')

def hyperlink_target_check(ws, df, col) -> dict:
    """
    Validare suplimentară: pentru fiecare celulă nenulă din df[col],
    citește cell.hyperlink.target din ws și verifică-l cu același regex.
    Returnează exact aceste câmpuri:
      - "% Hyperlink Pass Rate"
      - "Hyperlink Fail Count"
      - "Hyperlink Fail Example"
      - "Hyperlink Fail Example Cell Reference"
    """
    total = 0
    fails = 0
    first_fail = None

    # litera coloanei în foaie
    col_idx = df.columns.get_loc(col) + 1
    col_letter = get_column_letter(col_idx)

    for idx, text in (
        df[col]
            .dropna()
            .astype(str)
            .str.strip()
            .items()
    ):
        if not text:
            continue
        total += 1
        excel_row = idx + 2
        cell_ref = f"{col_letter}{excel_row}"
        cell = ws[cell_ref]
        link = cell.hyperlink.target if cell.hyperlink else None

        if not link or not VALID_URL_REGEX.match(link):
            fails += 1
            if first_fail is None:
                reason = "No hyperlink attached" if not link else f"Invalid hyperlink target ({link})"
                first_fail = (text, cell_ref, reason, link or "<no link>")

    pass_rate = f"{int((total - fails)/total*100)}%" if total else "0%"

    # Construim câmpurile pentru raport
    explanation = ""
    example = ""
    fail_cell_ref = ""
    if first_fail:
        txt, ref, reason, url = first_fail
        explanation    = reason
        example        = txt
        fail_cell_ref  = ref

    return report_check(
        f"{col} Hyperlink Target Check",
        "✅ Pass" if fails == 0 else "❌ Fail",
        Explanation=explanation,
        **{
            "% Hyperlink Pass Rate":                 pass_rate,
            "Hyperlink Fail Count":                  fails,
            "Hyperlink Fail Example":                example,
            "Hyperlink Fail Example Cell Reference": fail_cell_ref
        }
    )


def hyperlink_check(df: pd.DataFrame, col: str, ws=None, mapped_props=None) -> dict:
    """
    1) Text-based URL check
    2) Dacă text invalid și ws există: verifică cell.hyperlink.target
    Află litera reală a coloanei din ws folosind mapped_props.
    Returnează exact aceste câmpuri:
      - "% Pass Rate"
      - "Check Fail Count"
      - "Check Fail Example"
      - "Check Fail Example Cell Reference"
    """
    if ws is not None and mapped_props:
        header_name = mapped_props.get(col, col)
        col_letter = None
        for cell in ws[1]:
            if cell.value == header_name:
                col_letter = cell.column_letter
                break
        if col_letter is None:
            col_idx = df.columns.get_loc(col) + 1
            col_letter = get_column_letter(col_idx)
    else:
        col_idx = df.columns.get_loc(col) + 1
        col_letter = get_column_letter(col_idx)

    total = 0
    fails = 0
    first_fail = None

    for idx, val in df[col].items():
        text = str(val).strip()
        if not text:
            continue
        total += 1

        excel_row = idx + 2
        cell_ref = f"{col_letter}{excel_row}"

        ok = bool(VALID_URL_REGEX.match(text))

        reason = None
        if not ok and ws is not None:
            cell = ws[cell_ref]
            link = getattr(cell.hyperlink, "target", None)
            ok = bool(link and VALID_URL_REGEX.match(link))
            if not ok:
                reason = "Invalid hyperlink target" if link else "No hyperlink attached"
        elif not ok:
            reason = "Invalid URL format"

        if not ok:
            fails += 1
            if first_fail is None:
                first_fail = (text, cell_ref, reason)

    pass_rate = f"{int((total - fails)/total*100)}%" if total else "0%"
    explanation = ""
    example = ""
    fail_cell_ref = ""
    if first_fail:
        txt, ref, reason = first_fail
        explanation    = reason
        example        = txt
        fail_cell_ref  = ref

    return report_check(
        f"{col} Hyperlink Check",
        "✅ Pass" if fails == 0 else "❌ Fail",
        Explanation=explanation,
        **{
            "% Pass Rate":                         pass_rate,
            "Check Fail Count":                    fails,
            "Check Fail Example":                  example,
            "Check Fail Example Cell Reference":   fail_cell_ref
        }
    )



def group_i(df, mapped_props=None):
    """
    Optional Data - URL Field Checks
    Pentru fiecare din ["Product URL", "Product Video URL"]:
      A) dacă nu e mapat            → Skip "Not mapped"
      B) dacă e mapat dar lipsește → Skip "Mapped but column missing"
      C) filtrăm blank/header/duplicate
      D) dacă după filtrare nu rămâne niciun URL → Skip "No data present"
      E) altfel → apel hyperlink_check (text‐based + fallback on cell.hyperlink)
    """
    from openpyxl import Workbook
    import pandas as pd

    ws = None
    # Dacă ni se dă Workbook, îl transformăm în DataFrame
    if isinstance(df, Workbook):
        ws = df.active
        data = list(ws.values)
        df   = pd.DataFrame(data[1:], columns=data[0])

    # Redenumim header-ele conform mapped_props
    rename_map = {
        header: prop
        for prop, header in (mapped_props or {}).items()
        if header in df.columns
    }
    if rename_map:
        df = df.rename(columns=rename_map)

    url_cols = ["Product URL", "Product Video URL"]
    mapped   = set(mapped_props or [])
    results  = []

    skip_base = {
        "% Pass Rate":       "0%",
        "Check Fail Count":  0,
        "Check Fail Example": "",
        "Check Fail Example":"",
        "Check Fail Example Cell Reference": ""
    }

    for col in url_cols:
        check_name = f"{col} Hyperlink Check"

        # A) Not mapped
        if col not in mapped:
            results.append(report_check(
                check_name, "⏭️ Skip",
                Explanation="Not mapped",
                **skip_base
            ))
            continue

        # B) Mapped but missing
        if col not in df.columns:
            results.append(report_check(
                check_name, "⏭️ Skip",
                Explanation="Mapped but column missing",
                **skip_base
            ))
            continue

        # C) Pregătim seria curățată
        col_data = df[col]
        # Dacă avem duplicate headers, luăm prima coloană
        if isinstance(col_data, pd.DataFrame):
            col_data = col_data.iloc[:, 0]

        serie = col_data.fillna("").astype(str).str.strip()

        # D) Scoatem header-ul și duplicatele
        serie = serie[serie != col]
        serie = serie[~serie.duplicated()]

        # Verificăm dacă a rămas vreo valoare
        if serie.eq("").all():
            results.append(report_check(
                check_name, "⏭️ Skip",
                Explanation="No data present",
                **skip_base
            ))
            continue

        # E) Text‑based + fallback hyperlink
        filt = serie.to_frame(name=col)
        results.append(hyperlink_check(filt, col, ws=ws, mapped_props=mapped_props))

    return results





from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import pandas as pd

def group_j(df, mapped_props=None):
    """
    Product Name English - Mandatory Field - Character Limit Check (<=750 chars)
    → Skip dacă nu e mapat sau dacă lipsește coloana; altfel Pass/Fail.
    """
    col = "Product Name (English)"
    check_name = f"{col} Character Limit Check"
    max_len = 750
    mapped = set(mapped_props or [])

    # 1) Dacă există mapped_props dar col nu e în ele → Skip
    if mapped and col not in mapped:
        return [report_check(
            check_name,
            "⏭️ Skip",
            Explanation="Not mapped",
            **{"% Pass Rate": "0%", "Check Fail Count": 0, "Check Fail Example": "", "Check Fail Example Cell Reference": ""}
        )]

    # 2) Dacă e mapat dar lipsește coloana → Skip
    if mapped and col not in df.columns:
        return [report_check(
            check_name,
            "⏭️ Skip",
            Explanation="Mapped but column missing",
            **{"% Pass Rate": "0%", "Check Fail Count": 0, "Check Fail Example": "", "Check Fail Example Cell Reference": ""}
        )]

    # 3) Dacă primim un Workbook, îl transformăm în DataFrame
    if isinstance(df, Workbook):
        ws = df.active
        data = list(ws.values)
        df = pd.DataFrame(data[1:], columns=data[0])

    total = len(df)
    too_long = []

    # 4) Detectăm rândurile care depășesc limita
    for idx, val in df[col].fillna("").items():
        text = str(val).strip()
        length = len(text)
        if length > max_len:
            excel_row = idx + 2
            snippet = text[:25] + "..."
            too_long.append((excel_row, length, snippet))

    count = len(too_long)
    pct = f"{int((total - count) / total * 100)}%" if total else "100%"
    outcome = "✅ Pass" if count == 0 else "❌ Fail"
    details = {"% Pass Rate": pct, "Check Fail Count": count}

    # 5) Populăm Example și Explanation
    if count > 0:
        row_idx, length, snippet = too_long[0]
        col_letter = get_column_letter(df.columns.get_loc(col) + 1)
        details["Check Fail Example"] = snippet
        details["Check Fail Example Cell Reference"] = f"{col_letter}{row_idx}"
        details["Actual Length"] = length
        explanation = f"Over {max_len} chars"
    else:
        details["Check Fail Example"] = ""
        details["Check Fail Example Cell Reference"] = ""
        explanation = ""

    # 6) Construim raportul
    row = report_check(
        check_name,
        outcome,
        Explanation=explanation,
        **details
    )
    return [row]



from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd

def group_k(df, mapped_props=None):
    """
    Product Name (Local Language) - Optional Field - Character Limit Check (<=750 chars)
    → Skip dacă nu e mapat, Skip dacă e mapat dar lipsește coloana,
      altfel Pass/Fail după lungime.
    """
    prop       = "Product Name (Local Language)"
    check_name = f"{prop} Character Limit Check"
    max_len    = 750

    # 0) Extragem header‑ul din mapping
    header = (mapped_props or {}).get(prop)

    # 1) Not mapped → Skip
    if not header:
        return [report_check(
            check_name,
            "⏭️ Skip",
            Explanation="Not mapped",
            **{"Max Length": max_len, "Check Fail Count": 0, "Check Fail Example Cell Reference": ""}
        )]

    # 2) Dacă primim Workbook, îl transformăm în DataFrame
    if isinstance(df, Workbook):
        ws   = df.active
        data = list(ws.values)
        df   = pd.DataFrame(data[1:], columns=data[0])

    # 3) Mapped but column missing → Skip
    if header not in df.columns:
        return [report_check(
            check_name,
            "⏭️ Skip",
            Explanation="Mapped but column missing",
            **{"Max Length": max_len, "Check Fail Count": 0, "Check Fail Example Cell Reference": ""}
        )]

    # 4) Verificăm lungimea fiecărei celule din coloana mapată
    fail_count = 0
    first_ref  = ""
    for idx, val in df[header].fillna("").astype(str).items():
        text = val.strip()
        if len(text) > max_len:
            fail_count += 1
            if not first_ref:
                row_letter = get_column_letter(df.columns.get_loc(header) + 1)
                first_ref = f"{row_letter}{idx + 2}"

    # 5) Construim raportul
    outcome    = "✅ Pass" if fail_count == 0 else "❌ Fail"
    explanation= "" if fail_count == 0 else "Over 750 chars"
    return [report_check(
        check_name,
        outcome,
        Explanation=explanation,
        **{
            "Max Length": max_len,
            "Check Fail Count": fail_count,
            "Check Fail Example Cell Reference": first_ref
        }
    )]

from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl import Workbook

def group_l(df, mapped_props=None):
    """
    Product Descriptions - Optional Fields - Character Limit Check (<=4000 chars)
    → Skip dacă nu e mapat, Skip dacă e mapat dar lipsește coloana,
      altfel Pass/Fail după lungimea textului și raportează Example,
      Cell Reference și Actual Length (numai la Fail).
    """
    props   = [
        "Product Description (English)",
        "Product Description (Local Language)"
    ]
    max_len = 4000
    results = []

    # 1) Dacă primim Workbook, transformăm în DataFrame
    if isinstance(df, Workbook):
        ws   = df.active
        data = list(ws.values)
        df   = pd.DataFrame(data[1:], columns=data[0])

    # 2) Copiem valorile din header-urile mapate în cele două coloane canonice
    for prop in props:
        header = (mapped_props or {}).get(prop)
        if header and header in df.columns:
            df[prop] = df[header]

    total = len(df)
    for prop in props:
        check_name = f"{prop} Character Limit Check"
        # definim common fără Actual Length
        common = {
            "Max Length": max_len,
            "% Pass Rate": "",
            "Check Fail Count": 0,
            "Check Fail Example": "",
            "Check Fail Example Cell Reference": ""
        }

        # a) Not mapped → Skip
        if (mapped_props or {}).get(prop) is None:
            common["% Pass Rate"] = "0%"
            results.append(report_check(
                check_name,
                "⏭️ Skip",
                Explanation="Not mapped",
                **common
            ))
            continue

        # b) Mapped but column missing → Skip
        if prop not in df.columns:
            common["% Pass Rate"] = "0%"
            results.append(report_check(
                check_name,
                "⏭️ Skip",
                Explanation="Mapped but column missing",
                **common
            ))
            continue

        # c) Strângem toate valorile care depășesc max_len
        too_long = []
        for idx, val in df[prop].fillna("").items():
            text   = str(val).strip()
            length = len(text)
            if length > max_len:
                excel_row  = idx + 2  # rând Excel consideră header
                snippet    = text[:25] + ("…" if length > 25 else "")
                col_letter = get_column_letter(df.columns.get_loc(prop) + 1)
                too_long.append((excel_row, snippet, col_letter, length))

        count   = len(too_long)
        pct     = f"{int((total - count) / total * 100)}%" if total else "0%"
        outcome = "✅ Pass" if count == 0 else "❌ Fail"

        # d) Completez câmpurile comune
        common["% Pass Rate"]      = pct
        common["Check Fail Count"] = count

        # e) La Fail adaug Actual Length
        if too_long:
            row, snippet, letter, length = too_long[0]
            common["Check Fail Example"]               = snippet
            common["Check Fail Example Cell Reference"] = f"{letter}{row}"
            common["Actual Length"]                    = length

        # f) Adaug raportul final
        results.append(report_check(
            check_name,
            outcome,
            Explanation="" if count == 0 else f"Over {max_len} chars",
            **common
        ))

    return results



def group_m(df, mapped_props=None):
    """
    Mandatory Data - Single ProductID Per Cell
    Verifică să nu existe virgule în fiecare celulă din coloana Product ID,
    doar dacă Product ID a fost mapat. Altfel: ⏭️ Skip.
    """
    col = "Product ID"
    check_name = "Single ProductID Per Cell"
    mapped = set(mapped_props or [])

    # 1) Dacă există mapping, dar Product ID nu e printre ele → Skip
    if mapped and col not in mapped:
        return [report_check(
            check_name,
            "⏭️ Skip",
            Explanation="Not mapped",
            **{"Check Fail Count": 0, "Check Fail Example": "", "Check Fail Example Cell Reference": ""}
        )]

    # 2) Dacă e mapat, dar lipsește coloana → Skip
    if mapped and col not in df.columns:
        return [report_check(
            check_name,
            "⏭️ Skip",
            Explanation="Mapped but column missing",
            **{"Check Fail Count": 0, "Check Fail Example": "", "Check Fail Example Cell Reference": ""}
        )]

    # 3) Altfel, rulăm logica originală
    total = len(df)
    fails = []
    # găsim virgulele
    for idx, val in df[col].fillna("").astype(str).items():
        if "," in val:
            row_num = idx + 2
            fails.append((row_num, val.strip()))

    count = len(fails)
    outcome = "✅ Pass" if count == 0 else "❌ Fail"
    details = {"Check Fail Count": count}

    if count:
        r, example = fails[0]
        col_letter = chr(65 + df.columns.get_loc(col))
        details["Check Fail Example"] = example
        details["Check Fail Example Cell Reference"] = f"{col_letter}{r}"
        explanation = "Multiple identifiers present"
    else:
        details["Check Fail Example"] = ""
        details["Check Fail Example Cell Reference"] = ""
        explanation = ""

    row = report_check(
        check_name,
        outcome,
        Explanation=explanation,
        **details
    )
    return [row]




# def group_n(df, extra_id_cols=None, mapped_props=None):
#     # 0) Rename all mapped headers în coloane “prop”
#     # mapped_props: dict cu cheie=prop, valoare=header din catalog
#     rename_map = {}
#     for prop, header in (mapped_props or {}).items():
#         if header in df.columns:
#             rename_map[header] = prop
#     if rename_map:
#         df = df.rename(columns=rename_map)

#     # restul codului rămâne identic…
#     mapped_props  = mapped_props or {}
#     extra_id_cols = extra_id_cols or []
#     base_props = ["SKU", "EAN", "UPC", "GTIN", "CTIN", "ASIN"]
#     extra_props = [c for c in (extra_id_cols or []) if c and c not in base_props]
#     all_props = base_props + extra_props

#     results = []
#     for prop in all_props:
#         check_name = f"{prop} Single Secondary Product Identifier Per Cell"
#         common = {"Check Fail Count": 0, "Check Fail Example": "", "Check Fail Example Cell Reference": ""}

#         # 1) Not mapped → Skip
#         if prop not in (mapped_props or {}):
#             results.append(report_check(check_name, "⏭️ Skip", Explanation="Not mapped", **common))
#             continue

#         # 2) Column should now be named exactly `prop`
#         if prop not in df.columns:
#             results.append(report_check(check_name, "⏭️ Skip", Explanation="Column missing", **common))
#             continue

#         # 3) date checks…
#         col = prop
#         vals = df[col].fillna("").astype(str).map(trim_val)
#         if vals.eq("").all().all():
#             results.append(report_check(check_name, "⏭️ Skip", Explanation="No data present", **common))
#             continue

#         # 4) verificare virgule…
#         fails = []
#         loc = df.columns.get_loc(col)
#         # dacă ai duplicate, ia doar primul index
#         if isinstance(loc, (list, tuple, np.ndarray)):
#             loc = loc[0]
#         letter = get_column_letter(loc + 1) 
#         for idx, v in vals.items():
#             if "," in v:
#                 fails.append((idx + 2, v.strip()))

#         common["Check Fail Count"] = len(fails)
#         if fails:
#             outcome, expl = "❌ Fail", "Multiple identifiers present"
#             row, example = fails[0]
#             common["Check Fail Example"] = example
#             common["Check Fail Example Cell Reference"] = f"{letter}{row}"
#         else:
#             outcome, expl = "✅ Pass", ""

#         results.append(report_check(check_name, outcome, Explanation=expl, **common))

#     return results

from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

def group_n(df, extra_id_cols=None, mapped_props=None):
    """
    Optional Data – Single Secondary Product Identifier Per Cell

    mapped_props poate fi:
      • { prop: header }    (prop→header)
      • { header: prop }    (header→prop)
    Funcția normalizează mapping-ul în {prop:header}, redenumește
    coloanele, apoi raportează Pass/Fail după prezența virgulelor.
    """

    # 0a) Transformăm extra_id_cols într-o listă plată de șiruri
    if extra_id_cols is None:
        extra_list = []
    else:
        extra_list = []
        for item in extra_id_cols:
            if isinstance(item, (list, tuple, set, np.ndarray, pd.Index, pd.Series)):
                for sub in item:
                    extra_list.append(str(sub))
            else:
                extra_list.append(str(item))

    # 0b) Normalizăm mapped_props → dict(str→str)
    raw = {}
    if mapped_props is not None:
        # dacă e pandas.Series, îl transformăm în dict
        if isinstance(mapped_props, pd.Series):
            mapped_props = mapped_props.to_dict()
        # dacă nu e dict, încercăm să-l convertim
        if not isinstance(mapped_props, dict):
            try:
                mapped_props = dict(mapped_props)
            except Exception:
                raise ValueError("mapped_props trebuie să fie dict-like de str→str")
        # forțăm toate cheile și valorile la șiruri
        raw = { str(k): str(v) for k, v in mapped_props.items() }

    base_props = ["SKU", "EAN", "UPC", "GTIN", "CTIN", "ASIN"]
    # permitted = lista de stringuri, fără set-uri
    allowed = base_props + extra_list

    # 1) Detectăm orientarea mapping-ului → dacă e nevoie, inversăm header↔prop
    any_key = any(k in allowed for k in raw.keys())
    any_val = any(v in allowed for v in raw.values())
    if not any_key and any_val:
        raw = { v: k for k, v in raw.items() }

    # 2) Redenumim coloanele din df după raw (header → prop)
    rename_map = {}
    for prop, hdr in raw.items():
        if hdr in df.columns and prop in base_props:
            rename_map[hdr] = prop
    if rename_map:
        df = df.rename(columns=rename_map)

    # 3) Construim lista finală de proprietăți
    extra_props = [c for c in extra_list if c and c not in base_props]
    all_props   = base_props + extra_props

    results = []
    for prop in all_props:
        check_name = f"{prop} Single Secondary Product Identifier Per Cell"
        common = {
            "Check Fail Count": 0,
            "Check Fail Example": "",
            "Check Fail Example Cell Reference": ""
        }

        # a) Skip dacă nu e mapat
        if prop not in raw:
            results.append(report_check(check_name, "⏭️ Skip",
                                        Explanation="Not mapped", **common))
            continue

        # b) Skip dacă lipsește coloana
        if prop not in df.columns:
            results.append(report_check(check_name, "⏭️ Skip",
                                        Explanation="Column missing", **common))
            continue

        # c) Preiau seria (dacă sunt duplicate, doar prima coloană)
        series = df[prop]
        if isinstance(series, pd.DataFrame):
            series = series.iloc[:, 0]

        # d) Convert la șir și strip()
        vals = series.fillna("").astype(str).str.strip()
        if vals.eq("").all():
            results.append(report_check(check_name, "⏭️ Skip",
                                        Explanation="No data present", **common))
            continue

        # e) Detectăm vectorial virgulele
        mask = vals.str.contains(",", na=False)
        fail_count = int(mask.sum())
        common["Check Fail Count"] = fail_count

        if fail_count:
            first_idx  = mask[mask].index[0]
            excel_row  = first_idx + 2
            example    = vals.iat[first_idx]
            loc = df.columns.get_loc(prop)
            if isinstance(loc, (list, tuple, np.ndarray, pd.Index)):
                loc = loc[0]
            col_letter = get_column_letter(loc + 1)

            common.update({
                "Check Fail Example": example,
                "Check Fail Example Cell Reference": f"{col_letter}{excel_row}"
            })
            outcome, expl = "❌ Fail", "Multiple identifiers present"
        else:
            outcome, expl = "✅ Pass", ""

        results.append(report_check(check_name, outcome,
                                    Explanation=expl, **common))

    return results



def group_o(df, mapped_props=None):
    """
    Optional Data - Category Data Checks
    Pentru 'Category' și 'Sub-Category':
    - ⏭️ Skip dacă nu e mapat sau coloana lipsește
    - ❌ Fail dacă:
        • vreun text are >75 caractere
        • sau conține '<' ori '>'
    - ✅ Pass altfel
    """
    props = ["Category", "Sub-Category"]
    mapped = set(mapped_props or [])
    results = []

    # 1) Dacă primim Workbook, transformăm în DataFrame
    if isinstance(df, Workbook):
        ws = df.active
        data = list(ws.values)
        df = pd.DataFrame(data[1:], columns=data[0])

    for prop in props:
        check_name = f"{prop} Data Checks"
        common = {
            "Check Fail Count": 0,
            "Check Fail Example": "",
            "Check Fail Example Cell Reference": ""
        }

        # 2) Not mapped → Skip
        if prop not in mapped:
            results.append(report_check(
                check_name,
                "⏭️ Skip",
                Explanation="Not mapped",
                **common
            ))
            continue

        
        # 3) Mapped but column missing → Skip
        col = prop
        if col not in df.columns:
            results.append(report_check(
                check_name,
                "⏭️ Skip",
                Explanation="Mapped but column missing",
                **common
            ))
            continue
        
        # 4) Column present but completely empty → Skip
        series = (
            df[col]
            .fillna("")         # înlocuiește NaN cu ""
            .astype(str)        # forțează string
            .map(trim_val)      # taie spații
        )
        if series.eq("").all():
            results.append(report_check(
                check_name,
                "⏭️ Skip",
                Explanation="No data present",
                **common
            ))
            continue

        # 4) Parcurgem valorile și raportăm prima eroare
        fails = []
        for idx, text in series.items():  # folosim series cu trim
            if len(text) > 75 or re.search(r"[<>]", text):
                fails.append((idx + 2, text))
                break

        common["Check Fail Count"] = len(fails)
        if fails:
            row, example = fails[0]

    # prinde exact tag-ul sau semnele < >
            m = re.search(r'(<[^>]+>|[<>])', example)
            if m:
                snippet = m.group(0) 
            else:
                snippet = example[:25]

            letter = get_column_letter(df.columns.get_loc(col) + 1)
            common["Check Fail Example"] = snippet
            common["Check Fail Example Cell Reference"] = f"{letter}{row}"
            explanation = "Too long (>75)" if len(example) > 75 else "Contains prohibited char"
            outcome = "❌ Fail"
        else:
            outcome, explanation = "✅ Pass", ""

        results.append(report_check(
            check_name,
            outcome,
            Explanation=explanation,
            **common
        ))

    return results

def add_error_sheets(wb, output):
    """
    Pentru fiecare grup din output["validation_group_order"],
    adaugă un sheet 'Errors – <Group>' cu toate erorile.
    """

    for group_name in output.get("validation_group_order", []):
        items = output.get(group_name)
        # Dacă nu e listă de checks, sărim peste
        if not isinstance(items, list):
            continue

        # Colectăm doar checks cu Fail
        errors = []
        for it in items:
            if it.get("Check Outcome", "").startswith("❌"):
                ref = (
                    it.get("Check Fail Example Cell Reference")
                    or it.get("Empty Cell References", "")
                )
                val   = it.get("Check Fail Example", "")
                cnt   = it.get("Check Fail Count", 0) or it.get("Empty Cell Count", 0)
                errors.append((it["Check Performed"], ref, val, cnt))

        if not errors:
            continue

        safe_name = f"Errors – {group_name}"[:31]
        # Dacă sheet-ul deja există, îl ștergem
        if safe_name in wb.sheetnames:
            del wb[safe_name]
        ws: Worksheet = wb.create_sheet(safe_name)

        # Scriem header-ul
        ws.append(["Check Performed", "Cell Reference", "Example Value", "Fail Count"])
        for perf, ref, example, cnt in errors:
            ws.append([perf, ref, example, cnt])



if __name__ == "__main__":
    output, detailed_path = validate_file("test_urls.xlsx")
    import json
    # print(json.dumps(output["Optional Data - URL Field Checks"], indent=2))
